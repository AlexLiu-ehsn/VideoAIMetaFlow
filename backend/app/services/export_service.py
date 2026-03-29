import csv
import io
import json
from uuid import UUID

from openpyxl import Workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.services.video_service import get_video_detail


def _format_time(seconds: float) -> str:
    """秒數轉 MM:SS 格式"""
    m = int(seconds // 60)
    s = int(seconds % 60)
    return f"{m}:{s:02d}"


async def _build_video_metadata(db: AsyncSession, video_id: UUID) -> dict | None:
    """組裝單一影片的完整 metadata"""
    video = await get_video_detail(db, video_id)
    if not video:
        return None

    tags_by_category = {}
    for vt in video.video_tags:
        cat_label = vt.tag.category.label
        if cat_label not in tags_by_category:
            tags_by_category[cat_label] = []
        tags_by_category[cat_label].append(vt.tag.label)

    return {
        "id": str(video.id),
        "filename": video.filename,
        "duration": video.duration,
        "width": video.width,
        "height": video.height,
        "fps": video.fps,
        "status": video.status,
        "summary": video.summary,
        "critique": video.critique,
        "tags": tags_by_category,
        "segments": [
            {
                "index": seg.segment_index,
                "start_time": _format_time(seg.start_time),
                "end_time": _format_time(seg.end_time),
                "title": seg.title,
                "description": seg.description,
                "visual_description": seg.visual_description,
                "audio_description": seg.audio_description,
            }
            for seg in video.segments
        ],
        "critique_annotations": [
            {
                "timestamp": _format_time(ann.timestamp),
                "end_time": _format_time(ann.end_time) if ann.end_time else None,
                "type": ann.type,
                "comment": ann.comment,
                "severity": ann.severity,
            }
            for ann in video.critique_annotations
        ],
        "analyzed_at": str(video.analyzed_at) if video.analyzed_at else None,
        "created_at": str(video.created_at),
    }


async def export_json(db: AsyncSession, video_ids: list[UUID]) -> bytes:
    """匯出 JSON 格式"""
    data = []
    for vid in video_ids:
        meta = await _build_video_metadata(db, vid)
        if meta:
            data.append(meta)

    result = data[0] if len(data) == 1 else data
    return json.dumps(result, ensure_ascii=False, indent=2).encode("utf-8")


async def export_csv(db: AsyncSession, video_ids: list[UUID]) -> bytes:
    """匯出 CSV 格式（扁平化：一行一個分段）"""
    output = io.StringIO()
    writer = csv.writer(output)

    # 標頭
    writer.writerow([
        "影片名稱", "時長", "解析度", "狀態", "摘要",
        "標籤", "分段序號", "開始時間", "結束時間",
        "分段標題", "分段描述", "畫面描述", "語音描述",
    ])

    for vid in video_ids:
        meta = await _build_video_metadata(db, vid)
        if not meta:
            continue

        tags_str = "; ".join(
            f"{cat}: {', '.join(tags)}"
            for cat, tags in meta["tags"].items()
        )
        resolution = f"{meta['width']}x{meta['height']}" if meta.get("width") else ""

        if meta["segments"]:
            for seg in meta["segments"]:
                writer.writerow([
                    meta["filename"],
                    meta.get("duration", ""),
                    resolution,
                    meta["status"],
                    meta.get("summary", ""),
                    tags_str,
                    seg["index"],
                    seg["start_time"],
                    seg["end_time"],
                    seg.get("title", ""),
                    seg["description"],
                    seg.get("visual_description", ""),
                    seg.get("audio_description", ""),
                ])
        else:
            writer.writerow([
                meta["filename"],
                meta.get("duration", ""),
                resolution,
                meta["status"],
                meta.get("summary", ""),
                tags_str,
                "", "", "", "", "", "", "",
            ])

    return output.getvalue().encode("utf-8-sig")


async def export_xlsx(db: AsyncSession, video_ids: list[UUID]) -> bytes:
    """匯出 XLSX 格式（多 sheet）"""
    wb = Workbook()

    # Sheet 1: 影片總覽
    ws_overview = wb.active
    ws_overview.title = "影片總覽"
    ws_overview.append(["影片名稱", "時長(秒)", "解析度", "狀態", "摘要", "整體評析", "分析時間"])

    # Sheet 2: 標籤
    ws_tags = wb.create_sheet("標籤")
    ws_tags.append(["影片名稱", "標籤分類", "標籤"])

    # Sheet 3: 時間軸分段
    ws_segments = wb.create_sheet("時間軸分段")
    ws_segments.append([
        "影片名稱", "序號", "開始時間", "結束時間",
        "標題", "描述", "畫面描述", "語音描述",
    ])

    # Sheet 4: 製作人評析
    ws_critique = wb.create_sheet("製作人評析")
    ws_critique.append(["影片名稱", "時間戳", "結束時間", "類型", "評論", "嚴重程度"])

    for vid in video_ids:
        meta = await _build_video_metadata(db, vid)
        if not meta:
            continue

        name = meta["filename"]
        resolution = f"{meta['width']}x{meta['height']}" if meta.get("width") else ""

        # 總覽
        ws_overview.append([
            name, meta.get("duration"), resolution, meta["status"],
            meta.get("summary"), meta.get("critique"), meta.get("analyzed_at"),
        ])

        # 標籤
        for cat, tags in meta["tags"].items():
            for tag in tags:
                ws_tags.append([name, cat, tag])

        # 分段
        for seg in meta["segments"]:
            ws_segments.append([
                name, seg["index"], seg["start_time"], seg["end_time"],
                seg.get("title"), seg["description"],
                seg.get("visual_description"), seg.get("audio_description"),
            ])

        # 評析標註
        for ann in meta["critique_annotations"]:
            ws_critique.append([
                name, ann["timestamp"], ann.get("end_time"),
                ann["type"], ann["comment"], ann.get("severity"),
            ])

    # 自動調整欄寬（簡易版）
    for ws in [ws_overview, ws_tags, ws_segments, ws_critique]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
